"""C4 試験成績書(Excel) — 정의(계획) + store 최신 머지 상태로 元請け 납품물을 만든다 (HANDOFF_V3 C4).

3시트: 要約(全件/OK/NG/ERROR/MISSING/未実行/消化率) + 明細(全件: 체크리스트·항목·状態·結果時刻·차이)
       + 差分明細(NG의 全差分: 어느 行이 어떻게 — As-Is↔To-Be, 납품·감사용).
- 全件 = **정의 기준 전체 항목**(未実行 가시화), 消化率 = 実施/全件(실시율).
- 明細은 全件 — 행마다 결과 출처(run_id)를 박아 **낡은 OK가 최신으로 위장하지 못하게**(req2). NG 행은
  状態=NG로 구분되어 NG明細을 겸한다. MISSING_ASIS/MISSING_TOBE는 NG와 별도 状態(정답없음 vs 값틀림, req4).
- 差分明細은 NG 줄별 차이를 전부 펼쳐 "어느 行이 現→新으로 어떻게 틀렸는지"를 단일 문서로 보인다(감사 설득력).

openpyxl은 C4 전용 의존 — 함수 안에서 지연 import(없으면 명시 에러). print 금지(Core는 파일만 생성).
"""

from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path

from .models import ComparisonResult, ComparisonStatus, ShellDefinition
from .store import ShellRecord

_NOT_RUN = "未実行"  # 정의엔 있으나 실적(체크포인트)에 없는 항목

# 要約 집계 순서(件数 행). 未実行은 정의−실적 차이로 별도 산출.
_SUMMARY_STATUSES = [
    ComparisonStatus.OK,
    ComparisonStatus.NG,
    ComparisonStatus.ERROR,
    ComparisonStatus.MISSING_ASIS,
    ComparisonStatus.MISSING_TOBE,
]


@dataclass
class _Row:
    """明細 한 행(= 試験項目 1건). result 없으면 未実行."""

    test_id: str
    test_name: str
    item: str  # 출력 라벨
    status: str  # ComparisonStatus 값 또는 "未実行"
    run_id: str  # 결과 출처(시각). 未実行은 "-"
    detail: str  # 차이/에러/사유


def build_rows(
    definitions: list[ShellDefinition], records: list[ShellRecord]
) -> list[_Row]:
    """정의(계획) 기준으로 全件明細 행을 만든다 — 실적(records)을 항목 키로 매칭, 없으면 未実行.

    항목 키 = (test_id, output_name). output_name은 orchestrator 규약(다중 출력이면 라벨, 단일이면 None).
    """
    index: dict[tuple[str, str | None], tuple[ComparisonResult, str | None]] = {}
    for rec in records:
        for r in rec.results:
            index[(r.shell_id, r.output_name)] = (r, rec.run_id)

    rows: list[_Row] = []
    for d in definitions:
        multi = len(d.outputs) > 1
        for out in d.outputs:
            key = (d.test_id, out.label if multi else None)
            found = index.get(key)
            if found is None:
                rows.append(_Row(d.test_id, d.test_name, out.label, _NOT_RUN, "-", _NOT_RUN))
            else:
                result, run_id = found
                rows.append(_Row(
                    d.test_id, d.test_name, out.label,
                    result.status.value, run_id or "-", _detail(result),
                ))
    return rows


def build_diff_rows(
    definitions: list[ShellDefinition], records: list[ShellRecord]
) -> list[tuple[str, str, str, int, str, str]]:
    """NG 항목의 줄별 차이를 전부 펼친 행(감사용: 어느 行이 어떻게 틀렸는지).

    반환 튜플 = (test_id, test_name, 항목라벨, 行번호, As-Is, To-Be). NG가 아니거나 diff_lines가
    없으면 행을 만들지 않는다(차분 없음). 키 규약은 build_rows와 동일((test_id, output_name)).
    """
    index: dict[tuple[str, str | None], ComparisonResult] = {}
    for rec in records:
        for r in rec.results:
            index[(r.shell_id, r.output_name)] = r

    rows: list[tuple[str, str, str, int, str, str]] = []
    for d in definitions:
        multi = len(d.outputs) > 1
        for out in d.outputs:
            r = index.get((d.test_id, out.label if multi else None))
            if r is not None and r.status == ComparisonStatus.NG:
                for line in r.diff_lines:
                    rows.append((d.test_id, d.test_name, out.label,
                                 line.line_number, line.asis_content, line.tobe_content))
    return rows


def _detail(r: ComparisonResult) -> str:
    """状態별 차이/사유 요약(明細 셀). NG는 건수+첫 차이, 그 외는 사유."""
    if r.status == ComparisonStatus.NG:
        n = len(r.diff_lines)
        if r.diff_lines:
            d = r.diff_lines[0]
            return f"{n}件差異 / L{d.line_number} 現:{d.asis_content} → 新:{d.tobe_content}"
        return f"{n}件差異"
    if r.status == ComparisonStatus.ERROR:
        return r.error_message or "エラー"
    if r.status == ComparisonStatus.MISSING_ASIS:
        return "正解(As-Is)ファイルなし"
    if r.status == ComparisonStatus.MISSING_TOBE:
        return "出力(To-Be)ファイルなし"
    if r.status == ComparisonStatus.OK:
        return "一致"
    return ""


def summarize(rows: list[_Row]) -> dict:
    """要約 집계: 全件·状態별 件数·未実行·実施·消化率(実施/全件)."""
    total = len(rows)
    counts = {s.value: 0 for s in _SUMMARY_STATUSES}
    not_run = 0
    for row in rows:
        if row.status == _NOT_RUN:
            not_run += 1
        else:
            counts[row.status] = counts.get(row.status, 0) + 1
    done = total - not_run
    rate = (done / total * 100.0) if total else 0.0
    return {"total": total, "counts": counts, "not_run": not_run, "done": done, "rate": rate}


def generate_evidence(
    definitions: list[ShellDefinition], records: list[ShellRecord], output_path: Path
) -> Path:
    """試験成績書 .xlsx를 output_path에 생성하고 경로를 반환한다(要約 + 明細 + 差分明細 3시트)."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font
    except ImportError as exc:  # pragma: no cover - 환경 의존
        raise RuntimeError(
            "Excel 출력에 openpyxl이 필요합니다(pip install openpyxl). C4 試験成績書 전용 의존."
        ) from exc

    rows = build_rows(definitions, records)
    s = summarize(rows)

    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    bold = Font(bold=True)

    # --- 要約 시트 ---
    ws = wb.active
    ws.title = "要約"
    ws.append(["試験結果一覧 — 要約"])
    ws["A1"].font = bold
    ws.append([])
    ws.append(["項目", "件数"])
    for c in ws[ws.max_row]:
        c.font = bold
    ws.append(["全件", s["total"]])
    ws.append(["OK", s["counts"][ComparisonStatus.OK.value]])
    ws.append(["NG", s["counts"][ComparisonStatus.NG.value]])
    ws.append(["ERROR", s["counts"][ComparisonStatus.ERROR.value]])
    ws.append(["MISSING_ASIS(正解なし)", s["counts"][ComparisonStatus.MISSING_ASIS.value]])
    ws.append(["MISSING_TOBE(出力なし)", s["counts"][ComparisonStatus.MISSING_TOBE.value]])
    ws.append([_NOT_RUN, s["not_run"]])
    ws.append(["実施(全件−未実行)", s["done"]])
    ws.append(["消化率(実施/全件)", f"{s['rate']:.1f}%"])

    # --- 明細 시트(全件) ---
    ws2 = wb.create_sheet("明細")
    header = ["チェックリスト", "試験名", "項目", "状態", "結果時刻", "差異内容"]
    ws2.append(header)
    for c in ws2[1]:
        c.font = bold
    for row in rows:
        ws2.append([row.test_id, row.test_name, row.item, row.status, row.run_id, row.detail])

    # --- 差分明細 시트(NG 全差分: 어느 行이 現→新으로 어떻게 — 납품·감사용) ---
    ws3 = wb.create_sheet("差分明細")
    ws3.append(["チェックリスト", "試験名", "項目", "行", "As-Is(現)", "To-Be(新)"])
    for c in ws3[1]:
        c.font = bold
    diff_rows = build_diff_rows(definitions, records)
    for (tid, tname, item, ln, asis, tobe) in diff_rows:
        ws3.append([tid, tname, item, f"L{ln}", asis, tobe])
    if not diff_rows:
        ws3.append(["（差分なし）"])

    wb.save(output_path)
    return output_path
