#!/usr/bin/env python3
"""정의 매핑표 → 공유용 엑셀 템플릿(.xlsx) 생성 (D-048).

`definition_template.csv`(정본 20칼럼)를 그대로 .xlsx로 굽되, ★**모든 셀을 '텍스트 서식'으로 잠가**
Excel 자동변환(선두 0 제거 `00100`→`100`, `0:6`→시간, 큰 수 지수표기)으로 코드·layout·normalize가
조용히 깨지는 걸 막는다. 팀원은 이 .xlsx를 Excel로 채워 **그대로 제출** → `tools/mapping_to_definition.py`가
.xlsx를 직접 읽어 변환한다(CSV 저장 단계 불요, `read_mapping_bytes`).

사용:
  python tools/make_xlsx_template.py                  # definition_template.xlsx 생성(루트)
  python tools/make_xlsx_template.py -o path/to.xlsx
"""

from __future__ import annotations

import argparse
import csv
import sys
from pathlib import Path

_REPO = Path(__file__).resolve().parent.parent
_SRC_CSV = _REPO / "definition_template.csv"
_NAVY = "1F2937"  # 먹네이비(DESIGN_TOKENS) — 헤더 강조


def build_workbook(csv_path: Path):
    """정본 CSV(헤더+골격행)를 텍스트 서식 .xlsx 워크북으로 굽는다(openpyxl 지연 import)."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    rows = list(csv.reader(csv_path.read_text(encoding="utf-8").splitlines()))
    if not rows:
        raise ValueError(f"빈 템플릿 CSV: {csv_path}")
    header = rows[0]

    wb = Workbook()
    ws = wb.active
    ws.title = "mapping"
    for r, line in enumerate(rows, start=1):
        for c, val in enumerate(line, start=1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.number_format = "@"  # ★텍스트 — 선두0·layout·normalize 원문 보존(자동변환 차단)
    for c in range(1, len(header) + 1):
        h = ws.cell(row=1, column=c)
        h.font = Font(bold=True, color="FFFFFF")
        h.fill = PatternFill("solid", fgColor=_NAVY)
        h.alignment = Alignment(horizontal="center")
        ws.column_dimensions[h.column_letter].width = max(12, len(str(header[c - 1])) + 2)
    ws.freeze_panes = "A2"  # 헤더 고정(스크롤해도 칼럼명 보임)
    return wb


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(
        description="정의 매핑표(CSV 정본) → 공유용 엑셀 템플릿(.xlsx, 텍스트 서식 잠금) 생성"
    )
    parser.add_argument("-o", "--output", default=str(_REPO / "definition_template.xlsx"),
                        help="出力 .xlsx パス（기본: definition_template.xlsx）")
    parser.add_argument("--source", default=str(_SRC_CSV), help="정본 CSV(기본: definition_template.csv)")
    args = parser.parse_args(argv)

    src = Path(args.source)
    if not src.is_file():
        print(f"정본 CSV를 찾을 수 없습니다: {src}", file=sys.stderr)
        return 1
    try:
        wb = build_workbook(src)
    except ImportError:
        print("xlsx 생성에는 openpyxl이 필요합니다(pip install openpyxl).", file=sys.stderr)
        return 1
    wb.save(args.output)
    print(f"엑셀 템플릿 생성: {args.output}（텍스트 서식 잠금 — 선두0·layout 보존）", file=sys.stderr)
    print("→ 팀원이 Excel로 채워 제출 → mapping_to_definition.py가 .xlsx 직접 변환.", file=sys.stderr)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
