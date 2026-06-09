#!/usr/bin/env python3
"""매핑표(Long CSV) → test_definition.yaml 생성기 (Phase 7, D-035).

수기 YAML을 피하려고, 고객이 **스프레드시트(CSV)** 로 셸-입출력 매핑을 채우면 그것을
`test_definition.yaml`로 변환한다. 풀스키마(다중 입력·다중 출력, D-033)를 표현하려면
한 셸이 입력 N·출력 M(가변 길이)이라 **납작한 1행/셸로는 부족**하므로 **Long 형식**을 쓴다:
**입출력 항목당 1행, `shell_id`로 그룹화.**

설치 준비용 **독립 CLI 도구**다(GUI 아님 — T7-3 경량화 유지). make_golden.py처럼 tools/에 둔다.
CSV 채움 → 변환 1회 → yaml → config의 `definition_file`이 가리킴 → 경량 GUI/CLI로 실행.

한 체크리스트가 입력 여러 개(예: A DB + B DB)를 읽어 출력(예: 병합된 C DB)을 내는 구조라,
**checklist를 1차 키로 묶는 Long 형식**(입출력 항목당 1행, checklist로 그룹화)을 쓴다.

CSV 열(대소문자·순서 무관, 빈 칸 허용 / 괄호=구 이름 호환):
  checklist      [필수] 체크리스트 번호(=검증 단위). 같은 값 행들이 한 체크리스트로 묶임(입력 N·출력 M).
                 (구 이름: shell_id 도 1차 키로 받음.)
  io             [필수] input | output  — 이 행이 입력인지 출력인지.  (구: kind)
  type           [필수] database | file | sam | vsam — 저장 형식.  (구: db_or_file)
                 sam/vsam=고정길이 파일(file로 컴파일). sam=순차→byte 기본(mask/normalize 시 record+layout),
                 vsam=키순→record+layout+key 필수(없으면 경고). 상세 D-047.
  shell          [선택] 이 체크리스트에서 실행되는 배치(잡) 경로. 한 번만 적어도 됨(빈 칸=동봉 stub, 데모).
                 (구 이름: program)
  shell_group    [선택] 업무 그룹 태그(예: 業務A). **디렉토리 아님** — 경로·env는 config batch.groups가 듦(D-040).
                 체크리스트당 1값(행마다 다르면 에러). 비면 미사용(하위호환). ★lint(멤버십)는 preflight가 함.
  table          [type=database면 필수] As-Is 데이터를 적재할/결과가 쓰일 테이블명.
  input          [입력행] 입력 파일명. 비우면 자동 생성(아래 규칙). io=output 행에선 빈 칸.
  as_is_output   [출력행] As-Is 정답 파일명(↔ to_be_output 과 비교). 비우면 To-Be 출력명과 같게 자동.
  to_be_output   [출력행] To-Be 출력/Export 파일명. 비우면 자동. io=input 행에선 빈 칸.
  timeout        [선택] 초(기본 60).
  setup          [선택] 입력 적재 전 1회 실행할 준비 SQL(.sql)/스크립트 경로(체크리스트당 1회). 마스터·시퀀스 리셋용.
  in_encoding    [입력/선택] 입력 적재 인코딩(미지정 시 config 전역).

  ── 출력별 비교 옵션(선택 / 출력 행에서 읽어 compare 블록으로 운반) ──
  비우면 byte(바이트 완전 일치) 기본. 셀 내부 다중값(ignore_columns·fixed_layout·normalize_rules)은 `;` 구분(CSV `,` 충돌 방지).
  compare_mode   byte | text | record (미지정=byte).
  key_columns    record 정렬·정합 키(컬럼명 또는 인덱스). DB 출력은 사실상 필수(없으면 행 순서 false-NG).  (구: key)
  encoding       출력 판정 인코딩(없으면 config 전역).   delimiter 필드 구분자(기본 ,).   has_header true/false.
  ignore_columns 무시할 컬럼(실행일자·시퀀스ID 등; `;` 구분).   tolerance 수치 허용오차.  (구: mask)
  fixed_layout   고정길이(SAM) "start:end;start:end;..".  (구: layout)
  normalize_rules 컬럼별 정규화 "COL:rule[:arg];.." (rule: date/num:N/nullblank/zeropad:N/trim).  (구: normalize)

  ── 항목별 격납 디렉토리(선택, 사장님 규격 #4·#7·#11 / D-036) ──
  비우면 config.yaml 공통 디렉토리를 쓴다(권장). 업무·셸별로 위치가 흩어질 때만 적는다(적으면 그 행이 config보다 우선).
  input_dir     [입력행]  As-Is 입력 격납 디렉토리(#4).   없으면 asis_input_dir.
  as_is_dir     [출력행]  As-Is 정답 격납 디렉토리(#7).    없으면 asis_output_dir.
  to_be_dir     [출력행]  To-Be 출력 격납 디렉토리(#11).   없으면 tobe_output_dir.
  ※ 파일입력의 To-Be 스테이징 위치/파일명(구 dest_dir/dest_name, #7-3/#7-4)은 매핑표에서 제거 —
    config.tobe_input_dir 폴백. 행별 지정이 필요해지면 손YAML(InputSpec.dest_dir/dest_name)로(코어는 유지).

**빈 파일명 자동 규칙**(D-035, 사용자 확정 — 내용은 수기, 파일명은 규칙):
  · 셸의 입력(또는 출력)이 1개면 `{shell_id}.csv`.
  · 여러 개면 `{shell_id}_{테이블명}.csv`(테이블 없으면 `{shell_id}_in{n}` / `{shell_id}_out{n}.csv`).
  · 정답(expected)이 비면 그 출력의 To-Be 이름과 동일(폴더가 달라 충돌 없음).
  · **이미 적힌 이름은 그대로 존중**(자동은 빈 칸만). 입력CSV·정답·DB export명은 우리가 정하는 이름이라 안전.

엄격 생성: 한 행이라도 오류면 전체 거부(부분 생성=전체검증 착시 차단). 생성 yaml은
load_definitions로 **round-trip 재파싱**해 깨진 정의를 막는다(D-031 §2·3 계승).

사용:
  python tools/mapping_to_definition.py mapping.csv -o test_definition.yaml
  python tools/mapping_to_definition.py mapping.csv          # stdout으로
"""

from __future__ import annotations

import argparse
import csv
import io
import sys
import tempfile
from pathlib import Path

import yaml

# tools/를 직접 실행해도 repo 루트의 src 패키지를 import할 수 있게 path에 루트를 추가한다.
sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

from src.config.definition import DefinitionError, load_definitions  # noqa: E402

# 1차 키 = checklist(체크리스트 번호). 같은 값 행들이 한 체크리스트로 묶임(입력 N·출력 M).
# 구형 호환: shell_id도 1차 키로 받음. 실행 배치는 shell(구형: program) 열.
# file·expected는 빈 칸이면 규칙으로 자동 채우므로 필수 아님(table은 DB의 사실이라 필수).
_KEY_COLS = ("checklist", "shell_id")  # 둘 중 하나는 있어야(checklist 우선)
# 필수 논리열(별칭쌍: 신 이름 우선, 구 이름 호환). io=입출력 구분, type=DB냐 파일이냐.
# ★type은 D-046에서 db_or_file을 supersede(sam/vsam 값 도입으로 'db냐 file이냐'가 부정확). db_or_file=구 별칭.
_REQUIRED_ANY = (("io", "kind"), ("type", "db_or_file"))
_VALID_KIND = ("input", "output")
_VALID_TYPE = ("database", "file", "sam", "vsam")  # D-047: sam/vsam = file 저장의 고정길이 하위형식
_FILE_FORMATS = ("sam", "vsam")  # storage=file로 컴파일, 비교방식은 형식에서 도출(_apply_format_compare)
_VALID_MODE = ("byte", "text", "record")
# 출력 행 → compare 블록 운반. {YAML 키: (신 이름, 구 이름…)}. 신 이름 우선·구 이름 호환. 빈 칸은 운반 안 함.
_COMPARE_ALIASES = {
    "key": ("key_columns", "key"),
    "encoding": ("encoding",),
    "mask": ("ignore_columns", "mask"),
    "tolerance": ("tolerance",),
    "layout": ("fixed_layout", "layout"),
    "delimiter": ("delimiter",),
    "normalize": ("normalize_rules", "normalize"),
}
# 배치 경로 미지정 시 쓰는 동봉 stub(데모 기본 — 실 운영은 program 열에 실 배치 경로).
_STUB = {"database": "stub_batch/run_batch_db.py", "file": "stub_batch/run_batch_file.py"}


def _get(row: dict, *names: str) -> str:
    """별칭 중 첫 비지 않은 값(신 이름 우선, 구 이름 호환). 없으면 ''."""
    for n in names:
        v = row.get(n)
        if v:
            return v
    return ""


def mapping_to_definition(csv_text: str) -> dict:
    """Long CSV 텍스트 → {ok, yaml, count, shells, errors}. 오류가 하나라도 있으면 ok=False·yaml=''.

    shells는 [{test_id, input_count, output_count}]. errors는 사람용 메시지(행 번호 포함) 목록.
    빈 파일명은 _autofill_names가 규칙으로 채운다(내용은 수기, 파일명은 규칙 — D-035).
    """
    csv_text = csv_text.lstrip("﻿")          # BOM 제거(utf-8-sig 아닌 디코드 대비)
    csv_text = _strip_leading_comments(csv_text)  # 선두 #주석/공백 줄 허용(SAMPLE 경고 헤더 등)
    reader = csv.DictReader(io.StringIO(csv_text))
    if reader.fieldnames is None:
        return _fail(["CSVが空です（ヘッダー行が必要）。"])
    cols = {(c or "").strip().lower() for c in reader.fieldnames}
    missing = [group[0] for group in _REQUIRED_ANY if not any(c in cols for c in group)]
    if missing:
        return _fail([f"必須列がありません: {', '.join(missing)}"])
    if not any(k in cols for k in _KEY_COLS):
        return _fail([f"必須列がありません: checklist（または shell_id）"])

    errors: list[str] = []
    warnings: list[str] = []              # 비치명 경고(생성은 됨) — CLI/GUI가 사용자에게 표시
    order: list[str] = []                 # 셸 순서 보존
    shells: dict[str, dict] = {}

    for n, raw in enumerate(reader, start=2):  # 2 = 헤더 다음 첫 데이터 행(사람 기준 행 번호)
        row = {(k or "").strip().lower(): (v or "").strip() for k, v in raw.items()}
        sid = row.get("checklist") or row.get("shell_id") or ""  # checklist(1차 키) 우선, 구형 shell_id 호환
        kind = _get(row, "io", "kind").lower()        # io(구: kind) = input | output
        itype = _get(row, "type", "db_or_file").lower()  # type(구: db_or_file) = database | file
        if not sid:
            errors.append(f"{n}行目: checklist が空です。")
            continue
        if kind not in _VALID_KIND:
            errors.append(f"{n}行目[{sid}]: io は {_VALID_KIND} のいずれか（受領: '{kind}'）。")
            continue
        if itype not in _VALID_TYPE:
            errors.append(f"{n}行目[{sid}]: type は {_VALID_TYPE} のいずれか（受領: '{itype}'）。")
            continue

        sh = shells.get(sid)
        if sh is None:
            sh = {"inputs": [], "outputs": [], "programs": set(), "groups": set(),
                  "test_name": "", "timeout": "", "setup": ""}
            shells[sid] = sh
            order.append(sid)
        prog = row.get("shell") or row.get("program")  # shell(실행 배치) 우선, 구형 program 호환
        if prog and ";" in prog:  # 1:N 셸 시퀀스(실행 보류) — 형식만 거부, CSV 좌표로(Q1=a)
            errors.append(
                f"{n}行目[{sid}]: shell に ';'(シーケンス)が含まれます — 1:N順次実行は未対応"
                f"(実連携後に対応予定)。単一シェルで記入してください。"
            )
            continue
        if prog:
            sh["programs"].add(prog)
        grp = row.get("shell_group")  # 업무 그룹 태그(체크리스트당 1값)
        if grp:
            sh["groups"].add(grp)
        if row.get("test_name") and not sh["test_name"]:
            sh["test_name"] = row["test_name"]
        if row.get("timeout") and not sh["timeout"]:
            sh["timeout"] = row["timeout"]
        if row.get("setup") and not sh["setup"]:  # 체크리스트당 1회 준비 SQL/스크립트
            sh["setup"] = row["setup"]

        table = row.get("table", "")

        storage = _storage_type(itype)  # sam/vsam → file(고정길이 파일 저장). database/file은 그대로.
        # ★신 칼럼명(As-Is/To-Be 통일) → 기존 YAML 필드로 방출(코어 무수정). 칼럼=사람용, YAML=코어 계약.
        if kind == "input":
            if storage == "database" and not table:
                errors.append(f"{n}行目[{sid}]: 入力=database には table が必要です。")
                continue
            spec = {"csv": row.get("input", ""), "type": storage}   # input=입력 파일명. 빈 칸이면 _autofill_names.
            if storage == "database":
                spec["table"] = table
            if row.get("input_dir"):       # input_dir(#4 As-Is 입력 격납) → YAML src_dir
                spec["src_dir"] = row["input_dir"]
            if row.get("in_encoding"):     # 입력 적재 인코딩 override(선택)
                spec["in_encoding"] = row["in_encoding"]
            # ★dest_dir/dest_name(To-Be 입력 스테이징, #7-3/#7-4)은 매핑 CSV 표면에서 제거(A) —
            #   코어는 유지하되 행별 지정은 안 받음. 파일입력 스테이징은 config.tobe_input_dir 폴백.
            sh["inputs"].append(spec)
        else:  # output
            if storage == "database" and not table:
                errors.append(f"{n}行目[{sid}]: 出力=database には table が必要です。")
                continue
            mode = row.get("compare_mode", "").lower()
            if mode and mode not in _VALID_MODE:
                errors.append(f"{n}行目[{sid}]: compare_mode は {_VALID_MODE} のいずれか（受領: '{mode}'）。")
                continue
            tobe_name = row.get("to_be_output", "")  # To-Be 출력명. 빈 칸이면 _autofill_names.
            spec = {"type": storage, "expected": row.get("as_is_output", "")}  # as_is_output=정답 파일명 → YAML expected
            if storage == "database":
                spec["table"] = table
                spec["export_as"] = tobe_name
            else:
                spec["file"] = tobe_name
            if row.get("as_is_dir"):       # as_is_dir(#7 정답 격납) → YAML expected_dir
                spec["expected_dir"] = row["as_is_dir"]
            if row.get("to_be_dir"):       # to_be_dir(#11 To-Be 출력 격납) → YAML tobe_dir
                spec["tobe_dir"] = row["to_be_dir"]
            cmp_block = _compare_block(row, mode)  # P0 §3: 출력별 비교 옵션
            cmp_block = _apply_format_compare(itype, cmp_block, sid, n, warnings)  # D-047: sam/vsam 비교방식 도출
            if cmp_block:
                spec["compare"] = cmp_block
            sh["outputs"].append(spec)

    # 셸 단위 검증(입력·출력 ≥1, program 일관) + 빈 파일명 자동 채움.
    for sid in order:
        sh = shells[sid]
        if not sh["inputs"]:
            errors.append(f"[{sid}]: 入力（kind=input）が1件もありません。")
        if not sh["outputs"]:
            errors.append(f"[{sid}]: 出力（kind=output）が1件もありません。")
        if len(sh["programs"]) > 1:
            errors.append(f"[{sid}]: program が行ごとに異なります: {sorted(sh['programs'])}")
        if len(sh["groups"]) > 1:
            errors.append(f"[{sid}]: shell_group が行ごとに異なります: {sorted(sh['groups'])}")
        if sh["inputs"] and sh["outputs"]:
            _autofill_names(sid, sh)

    if not order and not errors:
        errors.append("有効なシェル行がありません。")
    if errors:
        return _fail(errors)

    tests = [_emit_shell(sid, shells[sid]) for sid in order]
    text = yaml.safe_dump({"tests": tests}, allow_unicode=True, sort_keys=False)

    # round-trip: 생성한 yaml을 실제 로더로 다시 파싱(깨진 정의 차단).
    rt_err = _roundtrip_error(text)
    if rt_err:
        return _fail([f"生成YAMLの再パースに失敗（ツールのバグの可能性）: {rt_err}"])

    summary = [
        {"test_id": _pad(sid), "input_count": len(shells[sid]["inputs"]),
         "output_count": len(shells[sid]["outputs"])}
        for sid in order
    ]
    return {"ok": True, "yaml": text, "count": len(order), "shells": summary, "errors": [], "warnings": warnings}


def _compare_block(row: dict, mode: str) -> dict:
    """출력 행의 비교 옵션 열을 compare 블록 dict로(빈 칸 제외). mode·has_header 포함.

    셀 내부 다중값(mask·layout·normalize)은 CSV 단계에서 `;` 구분 그대로 운반한다(V3 C2).
    """
    block: dict = {}
    if mode:
        block["mode"] = mode
    for yaml_key, names in _COMPARE_ALIASES.items():
        v = _get(row, *names)
        if v:
            block[yaml_key] = v
    hh = row.get("has_header", "").lower()
    if hh in ("true", "1", "yes", "y"):
        block["has_header"] = True
    return block


def _storage_type(itype: str) -> str:
    """매핑 CSV의 type 값을 YAML 저장 타입(database|file)으로. sam/vsam은 고정길이 '파일'이라 file로 컴파일."""
    return "file" if itype in _FILE_FORMATS else itype


def _apply_format_compare(itype: str, block: dict, sid: str, n: int, warnings: list) -> dict:
    """sam/vsam(고정길이 파일 하위형식)의 비교방식을 형식에서 도출한다(D-047, D-039).

    - sam : 기본 byte 통짜(순차파일·순서 의미). record 명시 또는 ignore_columns/normalize_rules가
            있으면 record+layout 필드비교. 그땐 layout 없으면 경고.
    - vsam: record+layout+key 필수(키순 저장 → 순서·byte 비교 불가). key/layout 없으면 경고.
    database/file은 그대로(무변경). block은 _compare_block 산출물을 in-place 조정해 돌려준다.
    ★sam/vsam은 Option A — YAML에선 type:file로 풀리므로 이 도구에서만 형식을 알 수 있어 lint도 여기서 한다.
    """
    if itype not in _FILE_FORMATS:
        return block
    has_layout = bool(block.get("layout"))
    field_opts = bool(block.get("mask") or block.get("normalize"))
    if itype == "vsam":
        if block.get("mode") and block["mode"] != "record":
            warnings.append(f"{n}行目[{sid}]: vsam は record 比較固定です（compare_mode='{block['mode']}'は無視）。")
        block["mode"] = "record"
        block.setdefault("has_header", False)  # 고정길이 = 헤더 없음(필드는 layout·인덱스)
        if not block.get("key"):
            warnings.append(
                f"{n}行目[{sid}]: vsam に key_columns がありません — VSAMはキー順格納のため整列キー必須（無いと行順false-NG）。"
            )
        if not has_layout:
            warnings.append(f"{n}行目[{sid}]: vsam に fixed_layout がありません — 固定長レコードの分割基準が必要です。")
    else:  # sam
        if block.get("mode") == "record" or field_opts:
            block["mode"] = "record"
            block.setdefault("has_header", False)
            if not has_layout:
                warnings.append(
                    f"{n}行目[{sid}]: sam でフィールド比較(record/ignore_columns/normalize_rules)には fixed_layout が必要です。"
                )
        else:
            block["mode"] = "byte"        # 기본: 순차 고정길이 통짜 바이트(D-039 — layout은 mask/normalize 시에만)
            block.pop("layout", None)     # byte는 layout 미사용 → 死데이터 방지
    return block


def _name(sid: str, count: int, idx: int, table: str | None, role: str) -> str:
    """파일명 베이스: 항목 1개면 {sid}, 여러 개면 {sid}_{테이블명}(없으면 {sid}_{role}{idx})."""
    if count == 1:
        return sid
    if table:
        return f"{sid}_{table}"
    return f"{sid}_{role}{idx}"


def _autofill_names(sid: str, sh: dict) -> None:
    """빈 파일명(input.csv / output.export_as|file / expected)을 규칙으로 채운다(D-035).

    내용(테이블·타입·배치)은 수기, **파일명만** 자동. 이미 적힌 값은 건드리지 않는다.
    """
    n_in, n_out = len(sh["inputs"]), len(sh["outputs"])
    for i, sp in enumerate(sh["inputs"], start=1):
        if not sp.get("csv"):
            sp["csv"] = _name(sid, n_in, i, sp.get("table"), "in") + ".csv"
    for k, sp in enumerate(sh["outputs"], start=1):
        base = _name(sid, n_out, k, sp.get("table"), "out")
        if sp["type"] == "database":
            if not sp.get("export_as"):
                sp["export_as"] = base + ".csv"
            tobe = sp["export_as"]
        else:
            if not sp.get("file"):
                sp["file"] = base + ".csv"   # 확장자 미상이면 csv 기본(필요 시 file 칸에 직접 기입)
            tobe = sp["file"]
        if not sp.get("expected"):
            sp["expected"] = tobe            # 정답은 To-Be와 같은 이름(폴더가 달라 충돌 없음)


def _emit_shell(sid: str, sh: dict) -> dict:
    """셸 1건을 로더가 읽는 신형 정의 dict로(input.tables[] + outputs[])."""
    program = next(iter(sh["programs"]), "") or _STUB[sh["inputs"][0]["type"]]
    entry: dict = {"test_id": sid}
    if sh["test_name"]:
        entry["test_name"] = sh["test_name"]
    entry["input"] = {"tables": sh["inputs"]}
    execution: dict = {"shell_program": program}
    if sh["timeout"]:
        execution["timeout"] = int(sh["timeout"])
    if sh.get("setup"):
        execution["setup"] = sh["setup"]
    grp = next(iter(sh["groups"]), "")  # 업무 그룹 태그(있으면 운반; 일관성은 위에서 검증)
    if grp:
        execution["shell_group"] = grp
    entry["execution"] = execution
    entry["outputs"] = sh["outputs"]
    return entry


def _roundtrip_error(yaml_text: str) -> str | None:
    """생성 yaml을 임시 파일로 써서 load_definitions로 검증. 통과면 None, 실패면 메시지."""
    with tempfile.NamedTemporaryFile("w", suffix=".yaml", delete=False, encoding="utf-8") as tf:
        tf.write(yaml_text)
        p = Path(tf.name)
    try:
        load_definitions(p)
        return None
    except DefinitionError as exc:
        return str(exc)
    finally:
        p.unlink(missing_ok=True)


def _pad(value: str) -> str:
    """test_id 3자리 zero-pad(로더와 동일 규칙). 비숫자는 그대로."""
    t = str(value).strip()
    return t.zfill(3) if t.isdigit() else t


def _strip_leading_comments(text: str) -> str:
    """선두의 빈 줄·'#' 주석 줄을 제거한다(헤더 행 앞 SAMPLE 경고 등 허용).

    데이터 행은 '#'로 시작하지 않으므로(checklist 값) 선두만 본다 — 본문 영향 없음(하위호환).
    """
    lines = text.splitlines()
    i = 0
    while i < len(lines) and (not lines[i].strip() or lines[i].lstrip().startswith("#")):
        i += 1
    return "\n".join(lines[i:])


def _fail(errors: list[str]) -> dict:
    return {"ok": False, "yaml": "", "count": 0, "shells": [], "errors": errors, "warnings": []}


def _decode(data: bytes) -> str:
    """Excel 저장(CSV) 대비: utf-8-sig 우선, 실패 시 cp932(일본어 Excel)."""
    for enc in ("utf-8-sig", "cp932"):
        try:
            return data.decode(enc)
        except UnicodeDecodeError:
            continue
    return data.decode("utf-8", errors="replace")


def _xlsx_to_csv_text(data: bytes) -> str:
    """.xlsx(첫 시트) → CSV 텍스트. 공유용 엑셀 템플릿을 CSV 변환 단계 없이 그대로 수용.

    셀은 텍스트 서식 전제(make_xlsx_template가 잠가둠 — 선두 0·layout·normalize 보존). openpyxl 지연 import
    (試験成績書와 동일 의존, 없으면 명시 에러). None→빈칸, 숫자는 str화, 완전 빈 행은 스킵.
    """
    import io as _io
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise ValueError("xlsx 입력에는 openpyxl이 필요합니다(pip install openpyxl).") from exc
    wb = load_workbook(_io.BytesIO(data), read_only=True, data_only=True)
    ws = wb.active
    buf = io.StringIO()
    w = csv.writer(buf, lineterminator="\n")
    for row in ws.iter_rows(values_only=True):
        cells = ["" if v is None else str(v) for v in row]
        if any(c.strip() for c in cells):  # 엑셀 잉여(완전 빈) 행 제외
            w.writerow(cells)
    return buf.getvalue()


def read_mapping_bytes(data: bytes) -> str:
    """매핑 입력(CSV 또는 .xlsx) 바이트 → CSV 텍스트. .xlsx(zip 시그니처)면 첫 시트를 변환, 아니면 디코드.

    CLI·GUI 공통 진입 — 팀원이 엑셀 템플릿을 채워 그대로 제출해도(CSV 저장 단계 없이) 받는다.
    """
    return _xlsx_to_csv_text(data) if data[:4] == b"PK\x03\x04" else _decode(data)


def main(argv: list[str] | None = None) -> int:
    """CLI: 매핑 CSV → yaml. 오류면 stderr에 행별 메시지 출력 + 종료코드 1."""
    parser = argparse.ArgumentParser(description="Long形式マッピング表(CSV/xlsx) → test_definition.yaml 生成")
    parser.add_argument("csv", help="入力マッピング表(CSV または .xlsx, Long形式)")
    parser.add_argument("-o", "--output", help="出力 yaml パス（省略時は標準出力）")
    args = parser.parse_args(argv)

    csv_path = Path(args.csv)
    if not csv_path.is_file():
        print(f"入力ファイルが見つかりません: {csv_path}", file=sys.stderr)
        return 1

    result = mapping_to_definition(read_mapping_bytes(csv_path.read_bytes()))
    if not result["ok"]:
        print("変換に失敗しました:", file=sys.stderr)
        for e in result["errors"]:
            print(f"  ・{e}", file=sys.stderr)
        return 1

    if args.output:
        Path(args.output).write_text(result["yaml"], encoding="utf-8")
        print(f"生成しました: {args.output}（{result['count']} シェル）", file=sys.stderr)
        for s in result["shells"]:
            print(f"  {s['test_id']}: 入力{s['input_count']}・出力{s['output_count']}", file=sys.stderr)
    else:
        sys.stdout.write(result["yaml"])
    for w in result.get("warnings", []):  # 비치명 경고(생성은 됨) — 사용자가 보게 stderr로
        print(f"  ⚠ {w}", file=sys.stderr)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
